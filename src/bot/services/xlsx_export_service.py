"""XLSX export service for completed campaign results."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class XlsxExportService:
    """Service for building XLSX files with campaign export data."""

    @staticmethod
    def _normalize_cell_value(value: Any) -> Any:
        """Convert Python values into XLSX-friendly cell values."""
        if value is None:
            return ""

        if isinstance(value, datetime):
            return value.replace(tzinfo=None) if value.tzinfo else value

        if isinstance(value, date):
            return value

        return value

    @staticmethod
    def _sanitize_filename(value: str) -> str:
        """Convert arbitrary text into a filesystem-safe filename fragment."""
        normalized = re.sub(r"[^\w\-]+", "_", value.strip(), flags=re.UNICODE)
        normalized = normalized.strip("._")
        return normalized or "campaign"

    def build_export_file(
        self,
        campaign_id: int,
        campaign_title: str,
        header: list[str],
        rows: list[list[Any]],
    ) -> tuple[str, BytesIO, int]:
        """
        Build XLSX export and return filename, file buffer and written row count.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"

        worksheet.append(header)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            worksheet.append([self._normalize_cell_value(value) for value in row])

        for column_cells in worksheet.columns:
            column_index = column_cells[0].column
            max_length = 0
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 40)

        safe_title = self._sanitize_filename(campaign_title)
        filename = f"campaign_{campaign_id}_{safe_title}.xlsx"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        written_rows = len(rows)
        return filename, buffer, written_rows